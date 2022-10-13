#! /usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

row_id = 1
stu_total = []
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value
        
        ws.cell(row = row_id, column = 7).value = sum_v
        stu_total.append(sum_v)
    row_id += 1

stu_total.sort(reverse=True)
stu_nu = len(stu_total)
g_nu = { 'A' : 0, 'B' : 0, 'C' : 0 }
for score in stu_total:
    if stu_total.index(score) + stu_total.count(score) <= stu_nu * 0.3:
        g_nu['A'] += 1
    elif stu_total.index(score) + stu_total.count(score) <= stu_nu * 0.7:
        g_nu['B'] += 1
    else:
        g_nu['C'] += 1

row_id = 1
for row in ws:
    if row_id != 1:
        grade = 'F'
        score = ws.cell(row = row_id, column = 7).value
        
        rank = stu_total.index(score)
        rank += stu_total.count(score)

        if rank <= stu_nu * 0.3:
            grade = 'A0'
            if rank <= g_nu['A'] * 0.5:
                grade = 'A+'
        elif rank <= stu_nu * 0.7:
            grade = 'B0'
            if rank <= g_nu['A'] + g_nu['B'] * 0.5:
                grade = 'B+'
        else:
            grade = 'C0'
            if rank <= g_nu['A'] + g_nu['B'] + g_nu['C'] * 0.5:
                grade = 'C+'
        ws.cell(row = row_id, column = 8).value = grade
    row_id += 1

wb.save( "student.xlsx" )
